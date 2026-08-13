import csv
import json
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = ROOT / "content" / "news.xlsx"
SOURCE_CSV = ROOT / "content" / "news-template.csv"
OUTPUT = ROOT / "data" / "news.json"
FIELDS = [
    "id", "title", "date", "time", "category", "summary", "body",
    "source_name", "source_url", "relevance", "image", "alt_text", "status", "featured", "is_demo",
]


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def read_xlsx() -> list[dict[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(value) for value in rows[0]]
    missing = [field for field in FIELDS if field not in headers]
    if missing:
        raise ValueError("Fehlende Excel-Spalten: " + ", ".join(missing))
    return [
        {field: clean(row[headers.index(field)] if headers.index(field) < len(row) else "") for field in FIELDS}
        for row in rows[1:]
        if any(clean(value) for value in row)
    ]


def read_csv() -> list[dict[str, str]]:
    with SOURCE_CSV.open("r", encoding="utf-8-sig", newline="") as file:
        return [{field: clean(row.get(field)) for field in FIELDS} for row in csv.DictReader(file)]


def main() -> None:
    rows = read_xlsx() if SOURCE_XLSX.exists() else read_csv()
    published = [row for row in rows if row["status"].lower() == "published"]
    published.sort(key=lambda row: (row["date"], row["time"], row["id"]), reverse=True)
    OUTPUT.write_text(json.dumps(published, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Generated {OUTPUT} with {len(published)} published item(s).")


if __name__ == "__main__":
    main()

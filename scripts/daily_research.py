"""Collect daily aviation candidates and stage them for editorial review.

This script never publishes automatically. It only appends candidates with
status=review to the public editorial workbook; the Pages build displays
published rows only.
"""

import hashlib
import re
import urllib.request
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "content" / "news.xlsx"
FEEDS = [
    ("aeroTELEGRAPH", "https://www.aerotelegraph.com/feed"),
    ("aero.de", "https://www.aero.de/feed/"),
    ("Fraport Newsroom", "https://www.fraport.com/de/newsroom/pressemitteilung.rss.xml"),
    ("Google News FRA", "https://news.google.com/rss/search?q=Frankfurt+Airport+OR+FRA+aviation&hl=de&gl=DE&ceid=DE:de"),
    ("Google News Europa Movements", "https://news.google.com/rss/search?q=aircraft+diversion+OR+emergency+landing+Europe+aviation&hl=en&gl=US&ceid=US:en"),
]
FRA_TERMS = ("frankfurt", "frankfurt airport", "fra", "fraport", "rhein-main", "lufthansa")
EUROPE_TERMS = ("europe", "europa", "germany", "deutschland", "france", "italy", "spain", "uk", "united kingdom", "switzerland", "austria", "netherlands", "belgium", "poland", "norway", "sweden", "denmark", "portugal", "greece", "ireland")
MOVEMENT_TERMS = ("diversion", "diverted", "emergency landing", "notlandung", "zwischenlandung", "return to", "umkehr", "7700", "smoke", "rauch", "odour", "odor", "geruch", "incident", "zwischenfall", "special movement")
FIELDS = ["id", "title", "date", "time", "category", "summary", "body", "source_name", "source_url", "relevance", "image", "alt_text", "status", "featured", "is_demo"]


def text(element: ET.Element | None) -> str:
    return "" if element is None else " ".join("".join(element.itertext()).split())


def parse_feed(source_name: str, url: str) -> list[dict[str, str]]:
    request = urllib.request.Request(url, headers={"User-Agent": "AeroNewsFRA-research/1.0"})
    try:
        with urllib.request.urlopen(request, timeout=20) as response:
            root = ET.fromstring(response.read())
    except Exception as error:
        print(f"Feed skipped: {source_name}: {error}")
        return []

    entries = []
    for item in root.iter():
        if item.tag.rsplit("}", 1)[-1] not in ("item", "entry"):
            continue
        title = text(next((child for child in item if child.tag.rsplit("}", 1)[-1] == "title"), None))
        description = text(next((child for child in item if child.tag.rsplit("}", 1)[-1] in ("description", "summary", "content")), None))
        link_node = next((child for child in item if child.tag.rsplit("}", 1)[-1] == "link"), None)
        link = (link_node.attrib.get("href", "") if link_node is not None else "") or text(link_node)
        if title and link:
            entries.append({"title": title, "description": description, "link": link, "source": source_name})
    return entries


def relevant(entry: dict[str, str]) -> tuple[bool, str]:
    haystack = f"{entry['title']} {entry['description']}".lower()
    fra = any(term in haystack for term in FRA_TERMS)
    movement = any(term in haystack for term in MOVEMENT_TERMS)
    europe = any(term in haystack for term in EUROPE_TERMS)
    if fra:
        return True, "Frankfurt Airport" if not movement else "Special Movement"
    if europe and movement:
        return True, "Special Movement Europa"
    return False, ""


def row_id(url: str) -> str:
    return "auto-" + hashlib.sha1(url.encode("utf-8")).hexdigest()[:12]


def append_candidates() -> int:
    workbook = load_workbook(WORKBOOK_PATH)
    sheet = workbook.active
    headers = [str(cell.value or "") for cell in sheet[1]]
    positions = {name: headers.index(name) for name in FIELDS}
    existing = {str(sheet.cell(row=row, column=positions["id"] + 1).value or "") for row in range(2, sheet.max_row + 1)}
    candidates = []
    for source_name, feed_url in FEEDS:
        for entry in parse_feed(source_name, feed_url):
            include, category = relevant(entry)
            identifier = row_id(entry["link"])
            if include and identifier not in existing:
                candidates.append((identifier, entry, category))
                existing.add(identifier)
    today = datetime.now(timezone.utc).date().isoformat()
    for identifier, entry, category in candidates[:20]:
        values = {
            "id": identifier,
            "title": entry["title"],
            "date": today,
            "time": "",
            "category": category,
            "summary": "Automatisch gefundener Kandidat. Fakten, FRA-Bezug und Quelle vor Veröffentlichung prüfen.",
            "body": entry["description"],
            "source_name": entry["source"],
            "source_url": entry["link"],
            "relevance": "review",
            "image": "",
            "alt_text": "",
            "status": "review",
            "featured": "no",
            "is_demo": "no",
        }
        sheet.append([values[field] for field in FIELDS])
    workbook.save(WORKBOOK_PATH)
    return len(candidates[:20])


if __name__ == "__main__":
    print(f"Staged {append_candidates()} candidate(s) for review.")
